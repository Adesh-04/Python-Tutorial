import openpyxl as xl
from openpyxl.reader.excel import load_workbook

filename = input("Enter File Name: ")

def add_numbers(filename):

    wb =  xl.load_workbook(filename)
    sheet = wb['Sheet3']
    # cell = sheet['a1'] # or sheet.cell(1,1)
    row_number = sheet.max_row # gives total rows

    for row in range(2, row_number +1):
        cell1 = sheet.cell(row,3)
        cell2 = sheet.cell(row,4)

        if cell1.value == None and cell2.value == None:
            continue

        elif cell1.value == None :
            cell1.value = 0
            addition = cell2.value
            new_cell = sheet.cell(row,5)
            new_cell.value = addition
            
        elif cell2.value == None :
            cell2.value = 0
            addition = cell1.value
            new_cell = sheet.cell(row,5)
            new_cell.value = addition
            
        else:     
            addition = cell1.value + cell2.value

            new_cell = sheet.cell(row,5)
            new_cell.value = addition
            if new_cell.value == "false":
                print("wrong")


    wb.save(filename)

add_numbers(filename)